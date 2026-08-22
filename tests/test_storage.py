"""저장 구조 테스트 (PRD v2.1 §9.3 파일명 규칙, §9.1 Manifest)."""

from __future__ import annotations

from datetime import date
from pathlib import Path

import pytest

from src.config_loader import load_topics
from src.models import AccessMode, PriorityLevel, Resource, ResourceStatus, SummaryBasis
from src.storage import Library, ManifestWriter, ResourceExporter, build_filename
from src.storage.manifest import MANIFEST_COLUMNS

from .conftest import CONFIG_DIR


@pytest.fixture(scope="module")
def topics():
    return load_topics(CONFIG_DIR / "topics.yaml")


def make_resource(**overrides) -> Resource:
    data = {
        "source_id": "kci",
        "title_original": "군사법원 판결의 증거능력 판단 기준",
        "title_ko": "군사법원 판결의 증거능력 판단 기준",
        "publisher": "한국형사·법무정책연구원",
        "publication_date": date(2026, 5, 12),
        "topic_primary": "01_군사법_군사사법",
        "doi": "10.1234/abcd.2026.05",
        "relevance_score": 82,
        "priority_level": PriorityLevel.P1,
        "access_mode": AccessMode.DOWNLOADED,
        "status": ResourceStatus.NEW,
        "summary_basis": SummaryBasis.FULLTEXT,
        "summary_ko": "**한줄 핵심**: 증거능력 판단 기준을 정리한다.",
    }
    data.update(overrides)
    return Resource(**data)


# ---------------------------------------------------------------------------
# §9.3 파일명 규칙
# ---------------------------------------------------------------------------

def test_filename_starts_with_yymmdd_download_date():
    """파일명은 다운로드일 YYMMDD 로 시작해야 합니다 (§9.3 필수)."""
    name = build_filename(
        downloaded_on=date(2026, 8, 22),
        source_id="KCI",
        title="군사재판절차 개선방안",
        extension=".pdf",
    )
    assert name.startswith("260822_")
    assert name == "260822_KCI_군사재판절차_개선방안.pdf"


def test_filename_uses_download_date_not_publication_date():
    """발행일이 아니라 다운로드일을 접두사로 씁니다 (§9.4)."""
    published = date(2026, 1, 5)
    downloaded = date(2026, 8, 22)
    name = build_filename(
        downloaded_on=downloaded, source_id="NKIS", title="공공부문 AI 법제연구", extension=".pdf"
    )
    assert name.startswith(downloaded.strftime("%y%m%d"))
    assert not name.startswith(published.strftime("%y%m%d"))


def test_filename_removes_windows_reserved_characters():
    name = build_filename(
        downloaded_on=date(2026, 8, 22),
        source_id="SSRN",
        title=r'AI/Military: Legal*Education?"<>|Study',
        extension=".pdf",
    )
    assert not any(ch in name for ch in r'\/:*?"<>|')


def test_filename_collapses_whitespace_to_underscore():
    name = build_filename(
        downloaded_on=date(2026, 8, 22),
        source_id="KCI",
        title="군사재판   절차   개선",
        extension=".pdf",
    )
    assert "  " not in name
    assert "군사재판_절차_개선" in name


def test_filename_respects_length_limit():
    name = build_filename(
        downloaded_on=date(2026, 8, 22),
        source_id="OPENALEX",
        title="가" * 400,
        extension=".pdf",
        max_length=160,
    )
    assert len(name) <= 160
    assert name.endswith(".pdf")


def test_filename_collision_suffix():
    """동일 파일명 충돌 시 해시 8자를 덧붙입니다 (§9.3)."""
    base = build_filename(
        downloaded_on=date(2026, 8, 22), source_id="KCI", title="군사재판절차", extension=".pdf"
    )
    with_hash = build_filename(
        downloaded_on=date(2026, 8, 22),
        source_id="KCI",
        title="군사재판절차",
        extension=".pdf",
        suffix_hash="a1b2c3d4",
    )
    assert base != with_hash
    assert "__a1b2c3d4.pdf" in with_hash


# ---------------------------------------------------------------------------
# §9.1 / §9.2 주제별 저장
# ---------------------------------------------------------------------------

def test_library_creates_all_topic_folders(project: Path, topics):
    Library(project / "data" / "library", topics)
    for topic in topics.topics:
        assert (project / "data" / "library" / topic.topic_id).is_dir()


def test_library_stores_into_topic_folder(project: Path, topics, sample_pdf: Path):
    library = Library(project / "data" / "library", topics)
    resource = make_resource(file_sha256="deadbeef" * 8)

    final = library.store(sample_pdf, resource, downloaded_on=date(2026, 8, 22))

    assert final.parent.name == "01_군사법_군사사법"
    assert final.name.startswith("260822_KCI_")
    assert final.exists()
    # staging 원본은 이동되었으므로 남아 있지 않아야 합니다.
    assert not sample_pdf.exists()


def test_library_stores_unknown_topic_into_unclassified(project: Path, topics, sample_pdf: Path):
    library = Library(project / "data" / "library", topics)
    resource = make_resource(topic_primary="존재하지_않는_주제")

    final = library.store(sample_pdf, resource)
    assert final.parent.name == "99_미분류_검토필요"


def test_library_handles_filename_collision(project: Path, topics, tmp_path: Path):
    from .conftest import minimal_pdf

    library = Library(project / "data" / "library", topics)
    first = tmp_path / "a.pdf"
    second = tmp_path / "b.pdf"
    first.write_bytes(minimal_pdf())
    second.write_bytes(minimal_pdf("다른 내용"))

    r1 = make_resource(file_sha256="1" * 64)
    r2 = make_resource(file_sha256="2" * 64)

    p1 = library.store(first, r1, downloaded_on=date(2026, 8, 22))
    p2 = library.store(second, r2, downloaded_on=date(2026, 8, 22))

    assert p1 != p2
    assert p1.exists() and p2.exists()


# ---------------------------------------------------------------------------
# Manifest
# ---------------------------------------------------------------------------

def test_manifest_written_to_dated_path(project: Path):
    writer = ManifestWriter(project / "data" / "manifests")
    paths = writer.write([make_resource()], collected_on=date(2026, 8, 22), run_id="run-1")

    assert paths.jsonl.exists()
    assert paths.csv.exists()
    assert paths.jsonl.parts[-3:] == ("2026", "08", "260822.jsonl")


def test_manifest_csv_has_required_columns(project: Path):
    writer = ManifestWriter(project / "data" / "manifests")
    paths = writer.write([make_resource()], collected_on=date(2026, 8, 22))

    header = paths.csv.read_text(encoding="utf-8-sig").splitlines()[0]
    for column in ("수집일", "제목", "주제", "최종저장경로", "DOI", "중요도", "요약근거수준"):
        assert column in header
    assert len(MANIFEST_COLUMNS) >= 20


def test_manifest_deduplicates_across_writes(project: Path):
    """같은 날 두 번 기록해도 CSV 에 중복 행이 쌓이지 않아야 합니다."""
    writer = ManifestWriter(project / "data" / "manifests")
    resource = make_resource()

    writer.write([resource], collected_on=date(2026, 8, 22))
    paths = writer.write([resource], collected_on=date(2026, 8, 22))

    rows = paths.csv.read_text(encoding="utf-8-sig").strip().splitlines()
    assert len(rows) == 2  # 헤더 + 1건


def test_manifest_records_query_provenance(project: Path):
    """어떤 검색어·사전 버전으로 발견했는지 남깁니다 (§3.4)."""
    writer = ManifestWriter(project / "data" / "manifests")
    resource = make_resource(
        discovered_by_query='"military court" OR court-martial',
        query_dictionary_version="2026.08.22-1",
        query_terms_expanded=["military court", "court-martial"],
    )
    paths = writer.write([resource], collected_on=date(2026, 8, 22))
    content = paths.csv.read_text(encoding="utf-8-sig")

    assert "court-martial" in content
    assert "2026.08.22-1" in content


# ---------------------------------------------------------------------------
# CSV / Excel 산출물
# ---------------------------------------------------------------------------

def test_exporter_writes_csv_with_bom(project: Path):
    exporter = ResourceExporter(
        project / "data" / "metadata" / "list.csv", project / "data" / "metadata" / "list.xlsx"
    )
    exporter.export([make_resource()], today=date(2026, 8, 22))

    raw = (project / "data" / "metadata" / "list.csv").read_bytes()
    assert raw.startswith(b"\xef\xbb\xbf"), "UTF-8-SIG BOM 이 필요합니다."


def test_exporter_creates_required_sheets(project: Path):
    from openpyxl import load_workbook

    excel = project / "data" / "metadata" / "list.xlsx"
    exporter = ResourceExporter(project / "data" / "metadata" / "list.csv", excel)

    resource = make_resource()
    resource.first_seen_at = None
    exporter.export([resource], today=date.today())

    wb = load_workbook(excel)
    for sheet in ("통합목록", "오늘수집", "P1_P2", "오류_검토필요"):
        assert sheet in wb.sheetnames
    # 월별 시트도 자동 생성
    assert any("목록" in name and "-" in name for name in wb.sheetnames)
