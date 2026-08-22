"""일자별 Manifest (PRD v2.1 §9.1, §9.2, §10.2).

수집일자별 "무엇을 수집했는지" 목록을 남깁니다. 실제 파일은 주제별 폴더에
한 번만 저장되므로, 날짜 기준 추적은 전적으로 Manifest 가 담당합니다.

경로: `data/manifests/YYYY/MM/YYMMDD.{jsonl,csv,xlsx}`
"""

from __future__ import annotations

import csv
import json
import logging
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from ..models import Resource, ResourceStatus

logger = logging.getLogger(__name__)

MANIFEST_COLUMNS = [
    "수집일",
    "신규수정구분",
    "제목",
    "원문제목",
    "출처",
    "발행기관",
    "발행일",
    "등록일",
    "수정일",
    "문서유형",
    "주제",
    "원문저장여부",
    "최종저장경로",
    "링크",
    "DOI",
    "공식ID",
    "파일해시",
    "텍스트해시",
    "중요도",
    "우선순위",
    "요약근거수준",
    "요약상태",
    "발견검색어",
    "검색어사전버전",
    "확장검색어",
    "라이선스",
    "오류검토필요",
]


@dataclass
class ManifestPaths:
    jsonl: Path
    csv: Path
    xlsx: Path


class ManifestWriter:
    """일자별 Manifest 를 JSONL / CSV / Excel 로 기록합니다."""

    def __init__(self, manifest_dir: str | Path):
        self.manifest_dir = Path(manifest_dir)

    # ------------------------------------------------------------------
    def paths_for(self, collected_on: date) -> ManifestPaths:
        folder = self.manifest_dir / f"{collected_on.year:04d}" / f"{collected_on.month:02d}"
        folder.mkdir(parents=True, exist_ok=True)
        stem = collected_on.strftime("%y%m%d")
        return ManifestPaths(
            jsonl=folder / f"{stem}.jsonl",
            csv=folder / f"{stem}.csv",
            xlsx=folder / f"{stem}.xlsx",
        )

    # ------------------------------------------------------------------
    def write(
        self,
        resources: list[Resource],
        *,
        collected_on: date | None = None,
        run_id: str = "",
    ) -> ManifestPaths:
        """자료 목록을 그날의 Manifest 에 추가 기록합니다."""
        collected_on = collected_on or date.today()
        paths = self.paths_for(collected_on)
        rows = [self.to_row(r, collected_on) for r in resources]

        self._append_jsonl(paths.jsonl, resources, collected_on, run_id)
        self._write_csv(paths.csv, rows)
        self._write_xlsx(paths.xlsx, rows, collected_on)

        logger.info("Manifest 기록 완료 (%d건): %s", len(rows), paths.csv)
        return paths

    # ------------------------------------------------------------------
    @staticmethod
    def to_row(resource: Resource, collected_on: date) -> dict[str, Any]:
        """Manifest 한 줄 (§10.2 필수 필드)."""
        return {
            "수집일": collected_on.isoformat(),
            "신규수정구분": resource.status.value,
            "제목": resource.best_title(),
            "원문제목": resource.title_original,
            "출처": resource.source_id,
            "발행기관": resource.publisher,
            "발행일": resource.publication_date.isoformat() if resource.publication_date else "",
            "등록일": resource.source_registered_date.isoformat()
            if resource.source_registered_date
            else "",
            "수정일": resource.source_modified_date.isoformat()
            if resource.source_modified_date
            else "",
            "문서유형": resource.document_type,
            "주제": resource.topic_primary,
            "원문저장여부": "저장" if resource.file_path else "링크만",
            "최종저장경로": resource.file_path,
            "링크": resource.canonical_url(),
            "DOI": resource.doi,
            "공식ID": resource.official_identifier,
            "파일해시": resource.file_sha256,
            "텍스트해시": resource.text_sha256,
            "중요도": resource.relevance_score,
            "우선순위": resource.priority_level.value,
            "요약근거수준": resource.summary_basis.value,
            "요약상태": "생성" if resource.summary_ko else "미생성",
            "발견검색어": resource.discovered_by_query,
            "검색어사전버전": resource.query_dictionary_version,
            "확장검색어": ", ".join(resource.query_terms_expanded),
            "라이선스": resource.license or ("불명확" if resource.license_unknown else ""),
            "오류검토필요": resource.error_code or "",
        }

    # ------------------------------------------------------------------
    def _append_jsonl(
        self, path: Path, resources: list[Resource], collected_on: date, run_id: str
    ) -> None:
        """원본 기록 — 전체 필드를 그대로 남깁니다 (감사 추적용, §18.4)."""
        with path.open("a", encoding="utf-8") as f:
            for resource in resources:
                record = json.loads(resource.model_dump_json())
                record["_manifest"] = {
                    "collected_on": collected_on.isoformat(),
                    "run_id": run_id,
                    "written_at": datetime.now().isoformat(),
                }
                f.write(json.dumps(record, ensure_ascii=False) + "\n")

    def _write_csv(self, path: Path, rows: list[dict[str, Any]]) -> None:
        """CSV 는 그날 전체를 JSONL 기준으로 다시 씁니다(중복 방지)."""
        existing = self._read_jsonl_rows(path.with_suffix(".jsonl"))
        merged = existing or rows
        with path.open("w", encoding="utf-8-sig", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=MANIFEST_COLUMNS, extrasaction="ignore")
            writer.writeheader()
            writer.writerows(merged)

    def _write_xlsx(self, path: Path, rows: list[dict[str, Any]], collected_on: date) -> None:
        try:
            from openpyxl import Workbook  # noqa: PLC0415
        except ImportError:
            logger.debug("openpyxl 이 없어 Manifest Excel 생성을 건너뜁니다.")
            return

        merged = self._read_jsonl_rows(path.with_suffix(".jsonl")) or rows
        wb = Workbook()
        ws = wb.active
        ws.title = f"{collected_on.strftime('%y%m%d')} 수집"
        ws.append(MANIFEST_COLUMNS)
        for row in merged:
            ws.append([row.get(col, "") for col in MANIFEST_COLUMNS])
        _autosize(ws)
        wb.save(path)

    # ------------------------------------------------------------------
    def _read_jsonl_rows(self, jsonl_path: Path) -> list[dict[str, Any]]:
        """JSONL 을 다시 읽어 Manifest 행으로 변환합니다."""
        if not jsonl_path.exists():
            return []
        rows: list[dict[str, Any]] = []
        seen: set[str] = set()
        with jsonl_path.open("r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    record = json.loads(line)
                except json.JSONDecodeError:
                    continue
                resource_id = record.get("resource_id", "")
                if resource_id in seen:
                    continue
                seen.add(resource_id)
                collected_on = date.fromisoformat(
                    (record.get("_manifest") or {}).get("collected_on", date.today().isoformat())
                )
                try:
                    resource = Resource(**{k: v for k, v in record.items() if not k.startswith("_")})
                except Exception:
                    continue
                rows.append(self.to_row(resource, collected_on))
        return rows

    # ------------------------------------------------------------------
    def read(self, collected_on: date) -> list[dict[str, Any]]:
        """특정 일자의 Manifest 를 읽습니다."""
        return self._read_jsonl_rows(self.paths_for(collected_on).jsonl)


def _autosize(worksheet: Any, *, max_width: int = 60) -> None:
    """열 너비를 내용에 맞춰 조정합니다."""
    from openpyxl.utils import get_column_letter  # noqa: PLC0415

    for index, column in enumerate(worksheet.iter_cols(), start=1):
        longest = max((len(str(cell.value or "")) for cell in column), default=10)
        worksheet.column_dimensions[get_column_letter(index)].width = min(max_width, longest + 2)


def status_label(status: ResourceStatus) -> str:
    return {
        ResourceStatus.NEW: "신규",
        ResourceStatus.UPDATED: "수정",
        ResourceStatus.DUPLICATE: "중복",
        ResourceStatus.LINK_ONLY: "링크만",
        ResourceStatus.FAILED: "실패",
        ResourceStatus.QUARANTINED: "격리",
    }.get(status, status.value)
