"""CSV / Excel 산출물 (PRD v2.1 §17.3).

- `list_download_resources.csv` : 전체 누적 메타데이터 (UTF-8-SIG)
- `list_download_resources.xlsx`: `통합목록` / `오늘수집` / `P1_P2` / `오류_검토필요`
  시트와 필요 시 월별 시트

Excel 은 편의용 결과물이며 SQLite 가 원본(Source of Truth)입니다.
"""

from __future__ import annotations

import csv
import logging
from collections import defaultdict
from datetime import date
from pathlib import Path
from typing import Any

from ..models import PriorityLevel, Resource

logger = logging.getLogger(__name__)

COLUMNS = [
    "수집일",
    "상태",
    "제목",
    "원문제목",
    "저자",
    "발행기관",
    "학술지_시리즈",
    "발행일",
    "등록일",
    "수정일",
    "출처",
    "문서유형",
    "주제",
    "중요도",
    "우선순위",
    "원문접근",
    "저장경로",
    "링크",
    "OA링크",
    "DOI",
    "공식ID",
    "라이선스",
    "요약근거수준",
    "요약",
    "발견검색어",
    "검색어사전버전",
    "확장검색어",
    "파일해시",
    "오류코드",
]

SHEET_ALL = "통합목록"
SHEET_TODAY = "오늘수집"
SHEET_PRIORITY = "P1_P2"
SHEET_ERRORS = "오류_검토필요"


def to_row(resource: Resource) -> dict[str, Any]:
    collected = resource.first_seen_at.date() if resource.first_seen_at else date.today()
    return {
        "수집일": collected.isoformat(),
        "상태": resource.status.value,
        "제목": resource.best_title(),
        "원문제목": resource.title_original,
        "저자": ", ".join(resource.authors),
        "발행기관": resource.publisher,
        "학술지_시리즈": resource.journal_or_series,
        "발행일": resource.publication_date.isoformat() if resource.publication_date else "",
        "등록일": resource.source_registered_date.isoformat()
        if resource.source_registered_date
        else "",
        "수정일": resource.source_modified_date.isoformat()
        if resource.source_modified_date
        else "",
        "출처": resource.source_id,
        "문서유형": resource.document_type,
        "주제": resource.topic_primary,
        "중요도": resource.relevance_score,
        "우선순위": resource.priority_level.value,
        "원문접근": resource.access_mode.value,
        "저장경로": resource.file_path,
        "링크": resource.canonical_url(),
        "OA링크": resource.oa_url,
        "DOI": resource.doi,
        "공식ID": resource.official_identifier,
        "라이선스": resource.license or ("불명확" if resource.license_unknown else ""),
        "요약근거수준": resource.summary_basis.value,
        "요약": _flatten_summary(resource.summary_ko),
        "발견검색어": resource.discovered_by_query,
        "검색어사전버전": resource.query_dictionary_version,
        "확장검색어": ", ".join(resource.query_terms_expanded),
        "파일해시": resource.file_sha256,
        "오류코드": resource.error_code,
    }


class ResourceExporter:
    """누적 목록을 CSV / Excel 로 내보냅니다."""

    def __init__(self, csv_path: str | Path, excel_path: str | Path):
        self.csv_path = Path(csv_path)
        self.excel_path = Path(excel_path)
        self.csv_path.parent.mkdir(parents=True, exist_ok=True)
        self.excel_path.parent.mkdir(parents=True, exist_ok=True)

    # ------------------------------------------------------------------
    def export(self, resources: list[Resource], *, today: date | None = None) -> None:
        rows = [to_row(r) for r in resources]
        self.write_csv(rows)
        self.write_excel(rows, today=today or date.today())

    # ------------------------------------------------------------------
    def write_csv(self, rows: list[dict[str, Any]]) -> None:
        with self.csv_path.open("w", encoding="utf-8-sig", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=COLUMNS, extrasaction="ignore")
            writer.writeheader()
            writer.writerows(rows)
        logger.info("CSV 저장 완료 (%d건): %s", len(rows), self.csv_path)

    # ------------------------------------------------------------------
    def write_excel(
        self, rows: list[dict[str, Any]], *, today: date, monthly_sheets: bool = True
    ) -> None:
        try:
            from openpyxl import Workbook  # noqa: PLC0415
            from openpyxl.styles import Alignment, Font, PatternFill  # noqa: PLC0415
        except ImportError:
            logger.warning("openpyxl 이 없어 Excel 생성을 건너뜁니다.")
            return

        wb = Workbook()
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="1A73E8")

        def add_sheet(title: str, sheet_rows: list[dict[str, Any]], index: int | None = None) -> None:
            ws = wb.create_sheet(title=title[:31], index=index)
            ws.append(COLUMNS)
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(vertical="center")
            for row in sheet_rows:
                ws.append([row.get(col, "") for col in COLUMNS])
            ws.freeze_panes = "A2"
            _autosize(ws)

        # 기본 시트 제거 후 순서대로 생성
        wb.remove(wb.active)
        add_sheet(SHEET_ALL, rows)

        today_iso = today.isoformat()
        add_sheet(SHEET_TODAY, [r for r in rows if r.get("수집일") == today_iso])

        priority_rows = [
            r for r in rows if r.get("우선순위") in (PriorityLevel.P1.value, PriorityLevel.P2.value)
        ]
        add_sheet(SHEET_PRIORITY, priority_rows)

        error_rows = [r for r in rows if r.get("오류코드")]
        add_sheet(SHEET_ERRORS, error_rows)

        if monthly_sheets:
            by_month: dict[str, list[dict[str, Any]]] = defaultdict(list)
            for row in rows:
                collected = str(row.get("수집일", ""))
                if len(collected) >= 7:
                    by_month[collected[:7]].append(row)
            for month in sorted(by_month, reverse=True):
                add_sheet(f"{month} 목록", by_month[month])

        wb.save(self.excel_path)
        logger.info("Excel 저장 완료: %s", self.excel_path)


def _autosize(worksheet: Any, *, max_width: int = 50) -> None:
    from openpyxl.utils import get_column_letter  # noqa: PLC0415

    for index, column in enumerate(worksheet.iter_cols(), start=1):
        longest = max((len(str(cell.value or "")) for cell in column[:200]), default=10)
        worksheet.column_dimensions[get_column_letter(index)].width = min(max_width, longest + 2)


def _flatten_summary(summary: str, limit: int = 900) -> str:
    """Excel 셀에 들어가도록 요약을 한 줄로 압축합니다."""
    if not summary:
        return ""
    flattened = " ".join(summary.split())
    return flattened[:limit]
