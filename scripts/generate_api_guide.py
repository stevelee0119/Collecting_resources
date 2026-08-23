"""`docs/API_발급_연동_가이드.md` 생성기 (PRD v2.1 §5.1 — 필수 산출물).

Source Registry 에서 **사전 발급·승인이 필요한 모든 Connector 를 자동 식별**하여
비개발자도 따라갈 수 있는 발급 절차 문서를 만듭니다.

보안 원칙:
- 실제 API Key / Client Secret / Refresh Token 은 절대 기록하지 않습니다.
- 환경변수명만 기록합니다.

최신성 원칙:
- 발급 절차·쿼터·가격은 변경될 수 있으므로 각 항목에 `verified_at` 을 남기고,
  `PENDING_VERIFICATION` 인 항목은 구현 직전 공식 문서 재확인을 요구합니다.
"""

from __future__ import annotations

import sys
from datetime import date
from pathlib import Path
from typing import Any

PROJECT_ROOT = Path(__file__).resolve().parent.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from src.config_loader import Settings, load_settings  # noqa: E402
from src.models import AuthType  # noqa: E402

UNKNOWN = "공식 문서 확인 필요"

#: 소스별 보충 정보. 공식 문서로 확인되지 않은 항목은 UNKNOWN 으로 두고
#: 임의로 지어내지 않습니다 (URL 환각 금지).
SOURCE_NOTES: dict[str, dict[str, Any]] = {
    "kci": {
        "purpose_example": "군사법·형사법 분야 국내 학술논문 메타데이터의 정기 수집 및 사내 리서치 색인 구축",
        "account_required": "예 (KCI 회원가입)",
        "menu_path": (
            "KCI 포털 → Open API 목록(openApiList.kci) → 사용하려는 API 선택 후 이용 신청. "
            "활용방법은 openApiConnSamp.kci, 명세서는 openDataView.kci 에서 확인"
        ),
        "approval": "신청 후 운영기관 검토",
        "pricing": UNKNOWN,
        "scopes": "해당 없음 (API Key 방식)",
        "redirect_uri": "불필요",
        "test": "python main.py run --daily --source kci --dry-run",
        "renewal": UNKNOWN,
        "cautions": (
            "OAI-PMH(https://open.kci.go.kr/oai/request)는 별도 인증 없이 사용 가능한 것으로 "
            "안내되어 있어 이 경로를 우선 사용합니다. Open API 는 apiCode/key/title/author/pubiYr "
            "파라미터를 사용하며, 일자 범위가 아닌 발행연도(pubiYr) 단위 필터만 제공합니다. "
            "제공 API: articleSearch / articleDetail / referenceSearch / citation / citationDetail"
        ),
    },
    "riss": {
        "purpose_example": "국방·법률 주제 학술자료의 서지 메타데이터 수집 (원문 자동 다운로드 없음)",
        "account_required": "예",
        "menu_path": "RISS API 센터(apicenter/apiMain.do)에서 제공 API 확인 후 신청",
        "approval": "기관 검토·승인 필요",
        "pricing": UNKNOWN,
        "scopes": "해당 없음",
        "redirect_uri": "불필요",
        "test": "python main.py run --daily --source riss --dry-run",
        "renewal": UNKNOWN,
        "cautions": (
            "RISS API 센터가 공개하는 API 는 상호대차(ILL)·E-DDS 신청, Rinfo 통계, "
            "FRIC 소장자원 검색용이며 일반 학술 서지 검색 API 는 확인되지 않았습니다. "
            "따라서 이 시스템은 RISS 를 직접 조회하지 않고 LINK_ONLY 로만 사용합니다. "
            "화면상 '원문있음' 표시가 자동 다운로드 권한을 의미하지 않습니다."
        ),
    },
    "scienceon": {
        "purpose_example": "국내외 학술논문·연구보고서 메타데이터 및 오픈액세스 원문 위치 탐색",
        "account_required": "예 (KISTI 통합회원)",
        "menu_path": (
            "ScienceON → OpenAPI(por/oapi/openApi.do) 신청 → API Gateway"
            "(apigateway/api/main/mainForm.do)에서 client_id 와 ACCESS_TOKEN 확인"
        ),
        "approval": UNKNOWN,
        "pricing": UNKNOWN,
        "scopes": "API 별 이용조건 확인 필요",
        "redirect_uri": "불필요",
        "test": "python main.py run --daily --source scienceon --dry-run",
        "renewal": "토큰 만료일이 있으므로 API Gateway 에서 갱신 상태를 확인하십시오.",
        "cautions": (
            "인증에 client_id 와 token(ACCESS_TOKEN) 두 값이 모두 필요합니다 "
            "(SCIENCEON_CLIENT_ID, SCIENCEON_API_KEY). 요청 파라미터는 "
            "version/action/target/searchQuery/curPage/rowCount 이며 searchQuery 는 JSON 문자열입니다. "
            "AccessON 오픈액세스 검색은 별도 이용조건이 적용될 수 있습니다."
        ),
    },
    "nkis": {
        "purpose_example": "정부출연연구기관 연구보고서의 정기 수집 및 내부 리서치 아카이브 구축",
        "account_required": "예",
        "menu_path": "NKIS 회원가입 → Open API 활용신청 → 기관검토 → 인증키 발급",
        "approval": "기관 검토 후 승인",
        "pricing": UNKNOWN,
        "scopes": "해당 없음",
        "redirect_uri": "불필요",
        "test": "python main.py run --daily --source nkis --dry-run",
        "renewal": UNKNOWN,
        "cautions": "발급받은 엔드포인트를 config/sources.yaml 의 endpoint 에 입력해야 동작합니다.",
    },
    "prism": {
        "purpose_example": "중앙·지방정부 정책연구용역 보고서의 정기 수집",
        "account_required": "예 (공공데이터포털 회원가입)",
        "menu_path": (
            "공공데이터포털 → '행정안전부_정책연구 과제정보'(15080254) → 활용신청. "
            "오퍼레이션: getResearchList_v2 / getResearchDetail_v2 / pnnMetaData_v2"
        ),
        "approval": "자동승인 또는 기관 검토 (API 별 상이)",
        "pricing": "공공데이터포털 기준 무료(트래픽 한도 있음)",
        "scopes": "해당 없음",
        "redirect_uri": "불필요",
        "test": "python main.py run --daily --source prism --dry-run",
        "renewal": "공공데이터포털 마이페이지에서 활용기간 연장 신청",
        "cautions": (
            "일반 인증키가 Encoding/Decoding 두 형태로 발급되므로 구분해서 사용하십시오. "
            "응답 필드명은 데이터셋 상세페이지의 '출력결과' 표를 보고 "
            "config/sources.yaml 의 field_map 을 맞춰야 합니다."
        ),
    },
    "law_go_kr": {
        "purpose_example": "군사법·국방 관련 법령·행정규칙·판례의 신규 제정 및 개정 추적",
        "account_required": "예",
        "menu_path": "국가법령정보 공동활용 → 오픈API 신청",
        "approval": "신청 후 승인 (승인 시 OC 식별자 부여)",
        "pricing": UNKNOWN,
        "scopes": "해당 없음",
        "redirect_uri": "불필요",
        "test": "python main.py run --daily --source law_go_kr --dry-run",
        "renewal": UNKNOWN,
        "cautions": (
            "OC 는 API Key 가 아니라 사용자 식별자이지만 동일하게 .env 로 관리합니다. "
            "요청변수는 OC/target/type/query/display/page 이며, 판례 목록은 prncYd(선고일자 범위)를 "
            "지원합니다. target 코드: law(법령) admrul(행정규칙) prec(판례) expc(법령해석례) ordin(자치법규). "
            "동일 데이터가 공공데이터포털에도 등재되어 있습니다(예: 법제처_판례 목록 조회 15059269)."
        ),
    },
    "crossref": {
        "purpose_example": "DOI 메타데이터 정규화 및 중복 판별 (연구 목적)",
        "account_required": "아니오",
        "menu_path": "별도 신청 없이 사용 가능. 연락 이메일(mailto)을 제공하면 polite pool 을 사용합니다.",
        "approval": "불필요",
        "pricing": "무료 (Plus 서비스는 유료)",
        "scopes": "해당 없음",
        "redirect_uri": "불필요",
        "test": "python main.py run --daily --source crossref --dry-run",
        "renewal": "해당 없음",
        "cautions": "CONTACT_EMAIL 을 설정하지 않으면 공유 풀에서 속도가 제한될 수 있습니다.",
    },
    "openalex": {
        "purpose_example": "국제법·법률AI 분야 영문 연구 동향 탐색",
        "account_required": "예 (API Key 발급 필요)",
        "menu_path": "openalex.org 로그인 → Settings → API 에서 무료 키 발급",
        "approval": "불필요 (즉시 발급)",
        "pricing": (
            "무료 키에 일일 예산이 배정되고 초과분은 사용량 기반 과금. "
            "키가 없으면 시험용 크레딧 100건 소진 후 409 반환"
        ),
        "scopes": "해당 없음",
        "redirect_uri": "불필요",
        "test": "python main.py run --daily --source openalex --dry-run",
        "renewal": "openalex.org/settings/api 에서 키 재발급·폐기",
        "cautions": (
            "2026-02-13 부터 모든 요청에 API Key 가 필수가 되었고 polite pool 과 mailto 파라미터는 "
            "폐지되었습니다. 사용량 기반 과금이므로 rate_limit_rps 를 낮게 유지하고 "
            "max_items_per_source 로 호출량을 통제하십시오."
        ),
    },
    "semantic_scholar": {
        "purpose_example": "중요 논문의 인용관계 탐색 및 최신 영문 논문 보강",
        "account_required": "API Key 신청 시 필요",
        "menu_path": "Semantic Scholar API 페이지의 Key 요청 양식 제출",
        "approval": "검토 후 발급",
        "pricing": "무료 (키 없으면 공유 쿼터)",
        "scopes": "해당 없음",
        "redirect_uri": "불필요",
        "test": "python main.py run --daily --source semantic_scholar --dry-run",
        "renewal": UNKNOWN,
        "cautions": (
            "키 없는 요청은 모든 미인증 사용자가 하나의 공유 키를 나눠 쓰므로 429 가 잦습니다. "
            "개인 키를 받으면 전 엔드포인트에서 초당 1회가 보장됩니다. "
            "대량 조회에는 /graph/v1/paper/search/bulk 를 사용하십시오."
        ),
    },
    "core": {
        "purpose_example": "오픈액세스 논문 원문 확보",
        "account_required": "예",
        "menu_path": "CORE → Services → API → 계정 등록 후 API Key 발급",
        "approval": UNKNOWN,
        "pricing": "무료 티어 제공 (상업적 이용은 별도 조건)",
        "scopes": "해당 없음",
        "redirect_uri": "불필요",
        "test": "python main.py run --daily --source core --dry-run",
        "renewal": UNKNOWN,
        "cautions": (
            "v3 는 쿼리 파라미터가 아니라 Authorization: Bearer 헤더로 인증합니다. "
            "원문 이용조건은 개별 논문 라이선스를 따릅니다."
        ),
    },
    "unpaywall": {
        "purpose_example": "DOI 기반 합법적 오픈액세스 원문 위치 확인",
        "account_required": "아니오",
        "menu_path": "별도 신청 없이 사용 가능. 모든 요청에 email 파라미터가 필요합니다.",
        "approval": "불필요",
        "pricing": "무료 (대량 이용은 데이터 덤프 권장)",
        "scopes": "해당 없음",
        "redirect_uri": "불필요",
        "test": "python main.py run --daily --source crossref --dry-run",
        "renewal": "해당 없음",
        "cautions": "API Key 가 아니라 연락 이메일을 요구합니다. CONTACT_EMAIL 을 반드시 설정하십시오.",
    },
    "doaj": {
        "purpose_example": "검증된 오픈액세스 저널 논문 메타데이터 수집",
        "account_required": "아니오 (일부 기능은 계정 필요)",
        "menu_path": "별도 신청 없이 검색 API 사용 가능",
        "approval": "불필요",
        "pricing": "무료",
        "scopes": "해당 없음",
        "redirect_uri": "불필요",
        "test": "python main.py run --daily --source doaj --dry-run",
        "renewal": "해당 없음",
        "cautions": (
            "현재 버전은 v4 이며 /api 와 /api/v4 양쪽에서 제공됩니다(2024-06 전환). "
            "pageSize 최대값은 100 입니다."
        ),
    },
    "arxiv": {
        "purpose_example": "AI·법률AI 분야 프리프린트 수집",
        "account_required": "아니오",
        "menu_path": "별도 신청 없이 사용 가능",
        "approval": "불필요",
        "pricing": "무료",
        "scopes": "해당 없음",
        "redirect_uri": "불필요",
        "test": "python main.py run --daily --source arxiv --dry-run",
        "renewal": "해당 없음",
        "cautions": "이용약관상 요청 간 최소 간격을 지켜야 합니다(기본 rate_limit_rps=0.33).",
    },
    "ssrn": {
        "purpose_example": "법학·사회과학 working paper 의 공식 Abstract Page 연결",
        "account_required": "아니오 (자동수집 대상 아님)",
        "menu_path": "공개 API 를 사용하지 않습니다. 외부 학술 메타데이터로 발견 후 링크만 보존합니다.",
        "approval": "해당 없음",
        "pricing": "해당 없음",
        "scopes": "해당 없음",
        "redirect_uri": "불필요",
        "test": "python main.py doctor",
        "renewal": "해당 없음",
        "cautions": "로그인·CAPTCHA·지문 우회를 통한 자동 다운로드는 금지됩니다.",
    },
    "zenodo": {
        "purpose_example": "연구보고서·프리프린트 보강 수집",
        "account_required": "토큰 발급 시 필요",
        "menu_path": "Zenodo 로그인 → Applications → Personal access tokens",
        "approval": "즉시 발급",
        "pricing": "무료",
        "scopes": "deposit:read 등 (읽기 전용 권장)",
        "redirect_uri": "불필요",
        "test": "python main.py run --daily --source zenodo --dry-run",
        "renewal": "Applications 화면에서 토큰 폐기·재발급",
        "cautions": "토큰 없이도 공개 레코드 검색이 가능합니다.",
    },
}

GMAIL_SECTION = """
## 부록 A. Gmail 발송 인증 (일일 브리핑)

PRD §15.4 는 SMTP 앱 비밀번호보다 **Gmail API + OAuth 2.0** 을 우선 검토하도록 규정합니다.

### A-1. Gmail API (권장)

| 항목 | 내용 |
| --- | --- |
| 1. 서비스/기관명 | Google — Gmail API |
| 2. 인증방식 | OAuth 2.0 (사용자 동의 기반) |
| 3. 사전 발급 필요 | 예 — Google Cloud 프로젝트, OAuth Client, Refresh Token |
| 4. 공식 발급 페이지 | https://console.cloud.google.com |
| 5. 계정 생성 | 예 (Google 계정) |
| 6. 신청 경로 | Google Cloud Console → 프로젝트 생성 → API 및 서비스 → 라이브러리 → Gmail API 사용 설정 → OAuth 동의 화면 구성 → 사용자 인증 정보 → OAuth 클라이언트 ID(데스크톱 앱) 생성 |
| 7. 서비스 목적 예시 | 내부 리서치 수집 결과의 일일 브리핑 자동 발송 |
| 8. 승인 절차 | 내부 사용(테스트 사용자 등록) 시 별도 검수 불필요. 외부 게시 시 Google 검수 대상 |
| 9. 무료/유료·쿼터 | 무료 (발송량 쿼터는 Google 정책 준수) |
| 10. OAuth Scope | `https://www.googleapis.com/auth/gmail.send` (발송 전용 최소 권한) |
| 11. Redirect URI | 데스크톱 앱 흐름에서 `http://localhost` 계열 사용 |
| 12. 환경변수명 | `GMAIL_CLIENT_SECRET_FILE`, `GMAIL_TOKEN_FILE` |
| 13. 동작 확인 | `python main.py run --daily --source arxiv` 후 수신함 확인 |
| 14. 키 갱신·회수 | Google Cloud Console → 사용자 인증 정보에서 클라이언트 폐기, 계정 보안 설정에서 액세스 권한 철회 |
| 15. 주의사항 | `client_secret.json` 과 토큰 파일을 Git 에 커밋하지 마십시오. |
| 16. 공식 문서 확인일 | `VERIFY_DATE` |

`config/config.yaml` 에서 `notification.email.transport: "gmail_api"` 로 설정합니다.

### A-2. SMTP 앱 비밀번호 (대안)

| 항목 | 내용 |
| --- | --- |
| 인증방식 | SMTP + 앱 비밀번호 |
| 신청 경로 | Google 계정 → 보안 → 2단계 인증 사용 설정 → 앱 비밀번호 생성 |
| 환경변수명 | `DLRCIS_SMTP_PASSWORD`, `DLRCIS_SENDER_EMAIL`, `DLRCIS_RECEIVER_EMAIL` |
| 주의사항 | 2단계 인증이 켜져 있어야 앱 비밀번호를 만들 수 있습니다. 계정 비밀번호를 그대로 쓰지 마십시오. |

## 부록 B. LLM 요약 (선택)

| 항목 | 내용 |
| --- | --- |
| 서비스 | Anthropic Claude API |
| 인증방식 | API Key |
| 공식 발급 페이지 | https://console.anthropic.com |
| 신청 경로 | Console 로그인 → API Keys → Create Key |
| 환경변수명 | `ANTHROPIC_API_KEY` |
| 동작 확인 | `config/config.yaml` 의 `summarizer.provider` 를 `llm` 로 변경 후 수집 실행 |
| 주의사항 | 키 미설정 시 자동으로 추출식 요약으로 대체됩니다. 요약 대상 원문의 외부 전송 가능 여부를 조직 정책으로 먼저 확인하십시오. |
"""


#: 인증정보가 실제로 설정되었는지 확인할 때 함께 필요한 부가 환경변수
COMPANION_ENV_VARS: dict[str, list[str]] = {
    "scienceon": ["SCIENCEON_CLIENT_ID"],
}


def _action_status(source: Any, method: Any) -> tuple[str, str]:
    """소스별 조치 현황을 판정합니다.

    Returns:
        (상태 아이콘+라벨, 남은 조치 설명)
    """
    from src.config_loader import get_secret  # noqa: PLC0415

    needs_key = method.credential_required and method.credential_env_var
    has_key = bool(get_secret(method.credential_env_var)) if needs_key else True

    missing_companions = [
        v for v in COMPANION_ENV_VARS.get(source.source_id, []) if not get_secret(v)
    ]

    if not method.endpoint:
        if method.verification_status == "VERIFIED":
            # 공식적으로 해당 API 가 없다고 확인된 경우 (예: RISS, SSRN)
            return "➖ 대상 아님", "자동수집 대상이 아닙니다. 추가 조치 불필요."
        return (
            "⬜ 조치 필요",
            f"공식 문서({source.auth_docs_url})에서 상세주소를 확인해 "
            f"`config/sources.yaml` 의 `endpoint` 에 입력",
        )

    if not needs_key:
        return "✅ 완료", "별도 발급 없이 사용 가능"

    if has_key and not missing_companions:
        return "✅ 완료", f"`{method.credential_env_var}` 설정 확인됨"

    if has_key and missing_companions:
        return (
            "🟡 일부 완료",
            f"`{method.credential_env_var}` 는 설정됨. "
            f"추가 필요: {', '.join(f'`{v}`' for v in missing_companions)}",
        )

    return "⬜ 조치 필요", f"발급 후 `.env` 에 `{method.credential_env_var}` 설정"



APPENDIX_D = """
## 부록 D. 정기 수집을 GitHub Actions 로 돌리는 경우

`credential-check.yml` 은 **점검 전용**이며 자료를 수집하지 않습니다.
GitHub Actions 에서 정기 수집까지 돌리려면 아래를 고려하십시오.

### D-1. 먼저 판단해야 할 것

| 항목 | 확인 사항 |
| --- | --- |
| 수집 파일 보관 | Actions 러너는 **실행이 끝나면 사라집니다.** `data/library/` 를 artifact 로 올리거나 외부 저장소에 커밋·업로드해야 유지됩니다. |
| SQLite 상태 | 중복 판별·증분 수집은 `data/metadata/dlrcis.db` 의 이력에 의존합니다. DB 를 매 실행마다 복원·저장하지 않으면 **매번 최초 실행처럼 동작**합니다. |
| 저작권 | 수집한 원문을 공개 저장소에 커밋하면 재배포가 됩니다. 라이선스가 불명확한 자료는 커밋하지 마십시오 (PRD §18.5). |
| 실행 시간 | Actions 작업에는 시간 제한이 있습니다. 백필처럼 오래 걸리는 작업은 로컬·서버 실행이 적합합니다. |

> 위 제약 때문에 **정기 수집은 사내 서버나 개인 PC 에서 `.env` + OS 스케줄러로 운영**하고,
> GitHub Actions 는 점검·테스트 용도로만 쓰는 구성을 권장합니다.
> Windows 는 `scripts/setup_scheduler.bat` 로 등록할 수 있습니다.

### D-2. 그래도 Actions 로 돌린다면

`.github/workflows/daily-collection.yml` 을 만들고
`credential-check.yml` 의 `env:` 블록을 그대로 복사한 뒤 아래를 참고하십시오.

```yaml
on:
  schedule:
    # UTC 기준입니다. KST 07:30 = UTC 22:30 (전날)
    - cron: "30 22 * * *"
  workflow_dispatch:

jobs:
  collect:
    runs-on: ubuntu-latest
    env:
      # credential-check.yml 의 env 블록을 그대로 복사
    steps:
      - uses: actions/checkout@v4
      - uses: actions/setup-python@v5
        with:
          python-version: "3.11"
      - run: pip install -r requirements.txt

      # 이전 실행의 DB 를 복원해야 증분 수집·중복 판별이 동작합니다.
      - uses: actions/cache@v4
        with:
          path: data/metadata
          key: dlrcis-db-${{ github.run_id }}
          restore-keys: dlrcis-db-

      - run: python main.py run --daily

      - uses: actions/upload-artifact@v4
        with:
          name: 수집결과
          path: |
            data/manifests/
            data/metadata/list_download_resources.xlsx
```

> `data/library/` (원문 파일)는 위 예시에서 **업로드하지 않습니다.**
> 저작권·라이선스를 검토한 뒤 보관 위치를 직접 정하십시오.
"""

def _method_label(verified_method: str) -> str:
    """무엇을 근거로 확정했는지 (§18.4 감사가능성)."""
    return {
        "official_doc": "공식 문서 직접 열람",
        "web_search": "공식 문서 검색결과 기준",
        "operator_input": "운영자가 직접 입력 (공식 문서 대조 전)",
        "none": "미확인",
    }.get(verified_method or "none", verified_method)


def _auth_label(auth_type: AuthType) -> str:
    return {
        AuthType.NONE: "불필요",
        AuthType.API_KEY: "API Key 필요",
        AuthType.OAUTH2: "OAuth 2.0 필요",
        AuthType.SERVICE_ACCOUNT: "서비스계정 필요",
        AuthType.INSTITUTION_APPROVAL: "기관 승인 필요",
        AuthType.OTHER: "기타 인증 필요",
    }.get(auth_type, str(auth_type))


def _source_section(source: Any, notes: dict[str, Any], today: str) -> str:
    lines: list[str] = [f"### {source.name} (`{source.source_id}`)", ""]

    methods = []
    for method in source.access_methods:
        status, todo = _action_status(source, method)
        methods.append(f"- **조치 현황: {status}** — {todo}")
        env = f"`{method.credential_env_var}`" if method.credential_env_var else "없음"
        methods.append(
            f"- **{method.type}** — 인증: {_auth_label(method.auth_type)} / "
            f"사전 발급 필요: {'예' if method.credential_required else '아니오'} / "
            f"환경변수: {env} / 확인상태: `{method.verification_status}`"
        )
        if method.verified_source:
            methods.append(
                f"  - 확인근거: {method.verified_source} "
                f"({method.verified_at or '-'}, {_method_label(method.verified_method)})"
            )
    lines.extend(methods)
    lines.append("")

    endpoints = [m.endpoint for m in source.access_methods if m.endpoint]
    endpoint_text = "\n".join(f"  - `{e}`" for e in endpoints) if endpoints else "  - (미설정)"

    rows = [
        ("1. 서비스/기관명", source.name),
        ("2. 사용 API/인증방식", ", ".join(f"{m.type}({_auth_label(m.auth_type)})" for m in source.access_methods)),
        ("3. 사전 발급 필요 여부", "예" if any(m.credential_required for m in source.access_methods) else "아니오"),
        ("4. 공식 발급/신청 페이지", source.auth_docs_url or UNKNOWN),
        ("5. 계정 생성 필요 여부", notes.get("account_required", UNKNOWN)),
        ("6. 신청 메뉴 경로", notes.get("menu_path", UNKNOWN)),
        ("7. 서비스 목적 예시", notes.get("purpose_example", UNKNOWN)),
        ("8. 승인 절차", notes.get("approval", UNKNOWN)),
        ("9. 무료/유료 및 쿼터", notes.get("pricing", UNKNOWN)),
        ("10. OAuth Scope/권한", notes.get("scopes", "해당 없음")),
        ("11. Redirect URI 필요 여부", notes.get("redirect_uri", "불필요")),
        (
            "12. 환경변수명",
            ", ".join(
                f"`{m.credential_env_var}`" for m in source.access_methods if m.credential_env_var
            )
            or "없음",
        ),
        ("13. 동작 확인 방법", f"`{notes.get('test', 'python main.py doctor')}`"),
        ("14. 키 만료·갱신·회수", notes.get("renewal", UNKNOWN)),
        ("15. 이용약관·자동수집 주의사항", notes.get("cautions", source.notes or UNKNOWN)),
        (
            "16. 공식 문서 최종 확인일",
            source.terms_checked_at.isoformat() if source.terms_checked_at else today,
        ),
    ]

    lines.append("| 항목 | 내용 |")
    lines.append("| --- | --- |")
    for label, value in rows:
        lines.append(f"| {label} | {str(value).replace('|', '/')} |")

    lines.append("")
    lines.append("**설정된 엔드포인트**")
    lines.append(endpoint_text)
    lines.append("")
    lines.append(f"**수집 정책**: `download_policy: {source.download_policy.value}` / "
                 f"`robots_policy: {source.robots_policy}` / `rate_limit_rps: {source.rate_limit_rps}`")
    lines.append("")

    pending = [m for m in source.access_methods if m.verification_status != "VERIFIED"]
    if pending:
        lines.append(
            "> ⚠ **구현 직전 재확인 필요** — 아래 접근방식은 공식 문서로 확정되지 않았습니다: "
            + ", ".join(m.type for m in pending)
        )
        lines.append("")

    lines.append("---")
    lines.append("")
    return "\n".join(lines)


def build_markdown(settings: Settings) -> str:
    today = date.today().isoformat()
    registry = settings.sources

    needs_credentials = []
    keyless = []
    for source in registry.sources:
        if any(m.credential_required or m.auth_type != AuthType.NONE for m in source.access_methods):
            needs_credentials.append(source)
        else:
            keyless.append(source)

    header = f"""# API 발급·연동 가이드

> 이 문서는 `config/sources.yaml` 에서 **자동 생성**됩니다.
> 수정하려면 `config/sources.yaml` 을 고친 뒤 `python main.py api-guide` 를 다시 실행하십시오.
>
> - 생성일: {today}
> - Source Registry 버전: `{registry.registry_version}`
> - 대상 소스: 전체 {len(registry.sources)}개 (사전 발급·승인 필요 {len(needs_credentials)}개)

## 0. 조치 현황 한눈에 보기

> **평가 환경: `GENERATED_CONTEXT`**
>
> 아래 상태는 이 문서를 생성한 환경에서 환경변수가 **실제로 읽히는지** 확인한 결과입니다.
> GitHub Secrets 에 값을 넣으셨다면 그 값은 **Actions 워크플로 실행 중에만** 존재하므로,
> 로컬에서 생성한 문서에는 "조치 필요"로 표시됩니다. 실제 상태를 보려면
> `.github/workflows/credential-check.yml` 을 수동 실행하십시오 (§0-3).

STATUS_SUMMARY_PLACEHOLDER

### 상태 표시의 의미

| 표시 | 의미 |
| --- | --- |
| ✅ 완료 | 이 환경에서 바로 사용 가능 (인증 불필요이거나 인증정보가 읽힘) |
| 🟡 일부 완료 | 인증정보 일부만 설정됨 — 남은 값을 추가해야 동작 |
| ⬜ 조치 필요 | 발급·엔드포인트 입력이 필요하거나, 값이 이 환경에 전달되지 않음 |
| ➖ 대상 아님 | 공식적으로 자동수집 대상이 아님 (추가 조치 불필요) |

---

## 0-1. 인증정보는 3단계를 모두 통과해야 동작합니다

키를 발급받았다고 해서 프로그램이 쓸 수 있는 것은 아닙니다.
아래 3단계가 **모두** 충족되어야 해당 소스가 실제로 수집됩니다.

```
1단계 발급     기관에서 API Key / 승인 / 식별자를 받았는가
   ↓
2단계 저장     안전한 곳에 보관했는가 (.env / OS 환경변수 / GitHub Secrets)
   ↓
3단계 전달     실행 시점에 프로그램의 환경변수로 주입되는가   ← 여기서 자주 막힙니다
   ↓
4단계 엔드포인트  config/sources.yaml 의 endpoint 가 채워져 있는가
```

### 보관 위치별 전달 여부

| 보관 위치 | 실행 환경 | 전달되는가 | 필요한 것 |
| --- | --- | --- | --- |
| `.env` (프로젝트 루트) | 로컬 / 서버 | **전달됨** | 없음 (`python-dotenv` 가 자동 로드) |
| OS 환경변수 | 로컬 / 서버 | **전달됨** | `export` (Linux/macOS) 또는 `setx` (Windows) |
| GitHub **Secrets** | GitHub Actions | 워크플로가 `env:` 로 매핑해야 전달됨 | **워크플로 파일** |
| GitHub **Secrets** | 로컬 / 서버 | **전달 안 됨** | 별도로 `.env` 필요 |
| GitHub **Variables** | 모든 환경 | 권장하지 않음 | — (아래 경고) |

> ⚠ **GitHub Actions Variables 에 API Key 를 넣지 마십시오.**
> Variables 는 **평문으로 저장**되어 저장소 읽기 권한자가 볼 수 있고 워크플로 로그에도 남습니다.
> API Key·토큰·비밀번호는 **Secrets** 에 저장하는 것이 맞습니다.
> (이미 Variables 에 넣었던 값이 있다면 Secrets 로 옮긴 뒤 Variables 에서 삭제하고,
> 노출 가능성이 있으므로 **키를 재발급**하는 것이 안전합니다.)

---

## 0-2. GitHub Secrets 에 등록할 이름

프로그램은 아래 **정확한 이름**의 환경변수를 읽습니다.
Secrets 이름을 이와 동일하게 맞추면 워크플로에서 그대로 매핑할 수 있습니다.

SECRET_NAMES_PLACEHOLDER

> Secrets 는 설계상 **값을 다시 읽을 수 없습니다**(write-only). 등록 여부는
> 이름 목록으로만 확인되며, 실제 동작 여부는 워크플로를 실행해 확인해야 합니다.

---

## 0-3. Secrets 가 실제로 동작하는지 확인하는 방법

`.github/workflows/credential-check.yml` 워크플로를 **수동 실행**하면
Secrets 가 주입된 상태에서 점검이 수행됩니다.

```
GitHub 저장소 → Actions 탭 → "인증정보 점검" → Run workflow
```

실행 결과에서 확인할 수 있는 것:

- 각 소스가 `[OK]` 인지 `[SKIP]` 인지 (=인증정보가 실제로 읽혔는지)
- 이 문서가 Secrets 기준으로 재생성되어 **artifact 로 첨부**됨
- 키 값 자체는 출력되지 않습니다 (GitHub 이 자동 마스킹하며, 코드도 마스킹 처리)

> 워크플로는 `workflow_dispatch` 전용이라 **직접 실행할 때만** 동작합니다.
> 자동 수집 스케줄을 원하시면 §부록 D 를 참고하십시오.

---

## 0-2. 읽기 전 안내

이 가이드는 **비개발자도 인증정보를 직접 발급받아 시스템에 연결할 수 있도록** 작성되었습니다.

### 보안 원칙 (PRD §5.1)

1. 실제 API Key, Client Secret, Refresh Token 을 **이 문서에 기록하지 마십시오.**
2. 이 문서에는 **환경변수명만** 적습니다. 예: `KCI_API_KEY`
3. 실제 값은 프로젝트 루트의 `.env` (Git 제외) 또는 OS Secret Store 에 저장합니다.
4. `.env.example` 에는 변수명과 설명만 넣고 실제 값은 넣지 않습니다.

### 최신성 원칙

API 발급·인증 방식은 변경될 수 있습니다. **Connector 를 실제로 연결하기 직전에 공식 문서를 다시 확인**하십시오.
블로그·카페·개인 튜토리얼은 보조자료로만 사용하고, 발급절차·쿼터·가격·권한은 공식 문서를 기준으로 판단합니다.

각 항목의 **확인상태**와 **확인근거**는 다음을 뜻합니다.

| 값 | 의미 |
| --- | --- |
| `VERIFIED` | 공식 문서로 엔드포인트·인증방식을 확정함 (확인근거 URL 기재) |
| `PENDING_VERIFICATION` | 아직 확정되지 않음 — 운영 전 공식 문서 확인 필요 |
| 확인근거 · 공식 문서 직접 열람 | 공식 페이지를 직접 열어 확인 |
| 확인근거 · 공식 문서 검색결과 기준 | 공식 문서의 검색 결과로 확인 (원문 재확인 권장) |

**엔드포인트가 비어 있으면 시스템은 해당 소스의 자동수집을 시도하지 않고 건너뜁니다.**

### 발급 절차 요약

```
1) 아래 표에서 필요한 소스를 고른다
2) '공식 발급/신청 페이지' 로 이동해 계정을 만들고 활용 신청을 한다
3) 발급받은 값을 프로젝트 루트 .env 파일에 '환경변수명=값' 형태로 적는다
4) 필요하면 config/sources.yaml 의 endpoint 를 공식 문서 기준으로 채운다
5) python main.py doctor 로 인식 여부를 확인한다
```

## 1. 인증정보가 필요한 소스 요약

| 소스 | 인증방식 | 환경변수 | 조치 현황 | 남은 조치 |
| --- | --- | --- | --- | --- |
"""

    summary_rows = []
    for source in needs_credentials:
        for method in source.access_methods:
            if not (method.credential_required or method.auth_type != AuthType.NONE):
                continue
            status, todo = _action_status(source, method)
            summary_rows.append(
                f"| {source.name} (`{source.source_id}`) | {method.type} / {_auth_label(method.auth_type)} "
                f"| `{method.credential_env_var or '-'}` "
                f"| {status} | {todo} |"
            )

    keyless_rows = "\n".join(
        f"- **{s.name}** (`{s.source_id}`) — 별도 발급 없이 사용 가능"
        + (" (연락 이메일 `CONTACT_EMAIL` 권장)" if s.source_id in ("crossref", "openalex", "unpaywall") else "")
        for s in keyless
    ) or "- (없음)"

    # 조치 현황 요약
    done, partial, todo_list, na = [], [], [], []
    for source in registry.sources:
        for method in source.access_methods:
            status, todo = _action_status(source, method)
            label = f"{source.name} (`{source.source_id}` / {method.type})"
            if status.startswith("✅"):
                done.append(label)
            elif status.startswith("🟡"):
                partial.append(f"{label} — {todo}")
            elif status.startswith("➖"):
                na.append(label)
            else:
                todo_list.append(f"{label} — {todo}")

    def bullets(items: list[str]) -> str:
        return "\n".join(f"- {i}" for i in items) if items else "- (없음)"

    status_summary = "\n".join([
        f"**✅ 조치 완료 ({len(done)}건)** — 바로 사용 가능",
        "",
        bullets(done),
        "",
        f"**🟡 일부 완료 ({len(partial)}건)** — 남은 값 추가 필요",
        "",
        bullets(partial),
        "",
        f"**⬜ 추가 조치 필요 ({len(todo_list)}건)**",
        "",
        bullets(todo_list),
        "",
        f"**➖ 자동수집 대상 아님 ({len(na)}건)** — 조치 불필요",
        "",
        bullets(na),
    ])

    # 이 문서를 어떤 환경에서 생성했는지 명시합니다.
    import os  # noqa: PLC0415

    if os.environ.get("GITHUB_ACTIONS") == "true":
        context = (
            f"GitHub Actions (workflow: {os.environ.get('GITHUB_WORKFLOW', '?')}) "
            "— Secrets 가 주입된 상태의 실제 결과입니다."
        )
    else:
        context = (
            "로컬/서버 실행 — GitHub Secrets 는 이 환경에 전달되지 않으므로, "
            "Secrets 에만 값을 넣으셨다면 '조치 필요'로 표시됩니다."
        )

    # 등록해야 할 Secret/환경변수 이름표
    secret_rows = ["| 환경변수 이름 | 용도 | 지금 등록이 필요한가 |", "| --- | --- | --- |"]
    seen_names: set[str] = set()
    for source in registry.sources:
        for method in source.access_methods:
            var = method.credential_env_var
            if not var or var in seen_names:
                continue
            seen_names.add(var)
            if not method.endpoint:
                # 엔드포인트가 없으면 키가 있어도 호출하지 않습니다.
                required = "➖ 불필요 (엔드포인트 미확정)"
            elif method.credential_required:
                required = "✅ 필수"
            else:
                required = "선택 (있으면 쿼터·속도 유리)"
            secret_rows.append(f"| `{var}` | {source.name} {method.type} | {required} |")
    for var, purpose, required in (
        ("SCIENCEON_CLIENT_ID", "ScienceON client_id (token 과 함께 필요)", "✅ ScienceON 사용 시 필수"),
        ("DLRCIS_SENDER_EMAIL", "브리핑 발신 주소", "✅ 알림 사용 시 필수"),
        ("DLRCIS_RECEIVER_EMAIL", "브리핑 수신 주소", "✅ 알림 사용 시 필수"),
        ("DLRCIS_SMTP_PASSWORD", "Gmail 앱 비밀번호", "✅ SMTP 사용 시 필수"),
        ("ANTHROPIC_API_KEY", "LLM 요약 (선택)", "선택"),
    ):
        if var not in seen_names:
            seen_names.add(var)
            secret_rows.append(f"| `{var}` | {purpose} | {required} |")
    secret_names_table = "\n".join(secret_rows)

    body = [
        header.rstrip("\n")
        .replace("STATUS_SUMMARY_PLACEHOLDER", status_summary)
        .replace("GENERATED_CONTEXT", context)
        .replace("SECRET_NAMES_PLACEHOLDER", secret_names_table),
        "\n".join(summary_rows),
        "",
        "## 2. 별도 발급 없이 사용 가능한 소스",
        "",
        keyless_rows,
        "",
        "> Crossref·OpenAlex 는 연락 이메일을 제공하면 polite pool 로 안정적인 응답을 받습니다.",
        "> Unpaywall 은 모든 요청에 이메일이 **필수**입니다.",
        "",
        "## 3. 소스별 상세 발급 절차",
        "",
    ]

    for source in registry.sources:
        body.append(_source_section(source, SOURCE_NOTES.get(source.source_id, {}), today))

    body.append(GMAIL_SECTION.replace("VERIFY_DATE", today))
    body.append("")
    body.append("## 부록 C. 환경변수 전체 목록")
    body.append("")
    body.append("| 환경변수 | 용도 | 필수 여부 |")
    body.append("| --- | --- | --- |")

    seen: set[str] = set()
    for source in registry.sources:
        for method in source.access_methods:
            var = method.credential_env_var
            if not var or var in seen:
                continue
            seen.add(var)
            body.append(
                f"| `{var}` | {source.name} {method.type} 인증 | "
                f"{'필수' if method.credential_required else '선택'} |"
            )
    for var, purpose, required in (
        ("CONTACT_EMAIL", "Crossref/OpenAlex polite pool, Unpaywall 필수 파라미터", "권장"),
        ("DLRCIS_SENDER_EMAIL", "브리핑 발신 주소", "필수"),
        ("DLRCIS_RECEIVER_EMAIL", "브리핑 수신 주소", "필수"),
        ("DLRCIS_SMTP_PASSWORD", "SMTP 앱 비밀번호", "SMTP 사용 시 필수"),
        ("GMAIL_CLIENT_SECRET_FILE", "Gmail OAuth 클라이언트 시크릿 파일 경로", "Gmail API 사용 시"),
        ("GMAIL_TOKEN_FILE", "Gmail OAuth 토큰 파일 경로", "Gmail API 사용 시"),
        ("ANTHROPIC_API_KEY", "LLM 요약(선택)", "선택"),
    ):
        if var in seen:
            continue
        seen.add(var)
        body.append(f"| `{var}` | {purpose} | {required} |")

    body.append("")
    body.append(APPENDIX_D)
    body.append("")
    body.append(f"> 마지막 생성: {today} · `python main.py api-guide` 로 재생성할 수 있습니다.")
    body.append("")
    return "\n".join(body)


def build_excel(settings: Settings, path: Path) -> Path | None:
    """비개발자 확인용 Excel 버전 (§5.1 선택 산출물)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        print("openpyxl 이 없어 Excel 생성을 건너뜁니다.")
        return None

    columns = [
        "소스ID", "서비스/기관명", "접근방식", "인증방식", "사전발급필요", "환경변수명",
        "공식발급페이지", "계정생성", "신청경로", "서비스목적예시", "승인절차",
        "무료/유료·쿼터", "Scope/권한", "RedirectURI", "동작확인", "키갱신·회수",
        "주의사항", "확인상태", "공식문서확인일", "엔드포인트",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "API 발급 가이드"
    ws.append(columns)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1A73E8")
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    today = date.today().isoformat()
    for source in settings.sources.sources:
        notes = SOURCE_NOTES.get(source.source_id, {})
        for method in source.access_methods:
            ws.append([
                source.source_id,
                source.name,
                method.type,
                _auth_label(method.auth_type),
                "예" if method.credential_required else "아니오",
                method.credential_env_var or "",
                source.auth_docs_url,
                notes.get("account_required", UNKNOWN),
                notes.get("menu_path", UNKNOWN),
                notes.get("purpose_example", UNKNOWN),
                notes.get("approval", UNKNOWN),
                notes.get("pricing", UNKNOWN),
                notes.get("scopes", "해당 없음"),
                notes.get("redirect_uri", "불필요"),
                notes.get("test", "python main.py doctor"),
                notes.get("renewal", UNKNOWN),
                notes.get("cautions", source.notes),
                method.verification_status,
                source.terms_checked_at.isoformat() if source.terms_checked_at else today,
                method.endpoint,
            ])

    from openpyxl.utils import get_column_letter

    for index, column in enumerate(ws.iter_cols(), start=1):
        longest = max((len(str(c.value or "")) for c in column), default=12)
        ws.column_dimensions[get_column_letter(index)].width = min(45, longest + 2)
    ws.freeze_panes = "A2"

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def generate(settings: Settings | None = None, *, include_excel: bool = False) -> list[Path]:
    """가이드 문서를 생성하고 경로 목록을 반환합니다."""
    settings = settings or load_settings()
    docs_dir = PROJECT_ROOT / "docs"
    docs_dir.mkdir(parents=True, exist_ok=True)

    md_path = docs_dir / "API_발급_연동_가이드.md"
    md_path.write_text(build_markdown(settings), encoding="utf-8")
    outputs = [md_path]

    if include_excel:
        xlsx = build_excel(settings, docs_dir / "API_발급_연동_가이드.xlsx")
        if xlsx:
            outputs.append(xlsx)
    return outputs


if __name__ == "__main__":
    for output in generate(include_excel="--excel" in sys.argv):
        print(f"생성: {output}")
